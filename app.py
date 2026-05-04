"""
TrackGTS Excel Charts Service — v2.0

Recibe un Excel del reporte de excesos de velocidad y le agrega 2 gráficos
de columnas con desglose por Conductor.

Cambios respecto a v1.0:
- Lee la hoja "Detalle 1" para extraer Conductor por exceso.
- Inserta columna "Conductor" en Resumen 1 entre Alias y Excesos.
- Si un vehículo tuvo varios conductores en la semana, genera una fila por
  combinación (vehículo, conductor) con Excesos y Velocidad Máxima
  recalculados desde Detalle 1.
- "Sin Conductor" cuando el campo Conductor viene vacío.
- Gráficos: categorías multi-nivel (Vehículo + Conductor).

Endpoints:
  GET  /           -> health check
  POST /add-charts -> recibe Excel base64, devuelve Excel base64 con gráficos

Autenticación: header `Authorization: Bearer <API_KEY>`
La API_KEY se configura como variable de entorno en Render.
"""
import os
import io
import base64
import logging
from collections import defaultdict, OrderedDict
from flask import Flask, request, jsonify
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableColumn

logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)

app = Flask(__name__)
API_KEY = os.environ.get('API_KEY', '')

# ---------- Constantes de layout (1-based, conocidas de TrackGTS) ----------
RESUMEN_HEADER_ROW = 5
RESUMEN_FIRST_DATA_ROW = 6
RESUMEN_COL_ALIAS = 1
RESUMEN_COL_EXCESOS_OLD = 2   # antes del insert
RESUMEN_COL_VELMAX_OLD = 6    # antes del insert

DETALLE_FIRST_DATA_ROW = 6
DETALLE_COL_ALIAS = 1
DETALLE_COL_CONDUCTOR = 2
DETALLE_COL_VELMAX = 12  # 'Velocidad Máxima (Km/h)'

SIN_CONDUCTOR = 'Sin Conductor'


# ---------- Auth ----------
def check_auth():
    """Valida el header Authorization: Bearer <API_KEY>."""
    if not API_KEY:
        log.warning('API_KEY no configurada — autenticación deshabilitada')
        return True
    auth = request.headers.get('Authorization', '')
    if not auth.startswith('Bearer '):
        return False
    return auth[7:].strip() == API_KEY


# ---------- Helpers ----------
def find_sheet(wb, keyword):
    """Encuentra una hoja cuyo nombre contenga keyword (case-insensitive)."""
    for name in wb.sheetnames:
        if keyword.lower() in name.lower():
            return name
    return None


def find_column_by_header(ws, header_row, keyword):
    """Devuelve el índice (1-based) de la columna cuyo header contenga keyword.

    Devuelve None si no se encuentra.
    """
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and keyword.lower() in str(val).lower():
            return col
    return None


def normalize_conductor(value):
    """Normaliza el campo Conductor: vacío/None -> 'Sin Conductor'."""
    if value is None:
        return SIN_CONDUCTOR
    s = str(value).strip()
    return s if s else SIN_CONDUCTOR


def aggregate_detalle(ws_detalle):
    """Recorre Detalle 1 y devuelve {alias: OrderedDict[conductor: {excesos, vel_max}]}.

    Mantiene el orden de primera aparición de los conductores para cada vehículo.
    """
    by_vehicle = defaultdict(OrderedDict)
    for r in range(DETALLE_FIRST_DATA_ROW, ws_detalle.max_row + 1):
        alias = ws_detalle.cell(row=r, column=DETALLE_COL_ALIAS).value
        if not alias:
            continue
        alias = str(alias).strip()
        conductor = normalize_conductor(
            ws_detalle.cell(row=r, column=DETALLE_COL_CONDUCTOR).value
        )
        vel_raw = ws_detalle.cell(row=r, column=DETALLE_COL_VELMAX).value
        try:
            vel = float(vel_raw) if vel_raw is not None and str(vel_raw).strip() != '' else 0.0
        except (ValueError, TypeError):
            vel = 0.0

        if conductor not in by_vehicle[alias]:
            by_vehicle[alias][conductor] = {'excesos': 0, 'vel_max': 0.0}
        by_vehicle[alias][conductor]['excesos'] += 1
        if vel > by_vehicle[alias][conductor]['vel_max']:
            by_vehicle[alias][conductor]['vel_max'] = vel
    return by_vehicle


def rebuild_resumen_with_conductor(ws_resumen, by_vehicle):
    """Inserta columna 'Conductor' en B y desglosa filas por conductor.

    Para vehículos con un solo conductor preserva los valores originales
    (Tiempo, Distancia, etc.). Para multi-conductor, recalcula Excesos y
    Velocidad Máxima desde Detalle 1; el resto de columnas quedan vacías.

    Devuelve la última fila con datos tras la modificación.
    """
    # 0. Capturar y eliminar tablas estructuradas existentes (se recrean al final
    #    con el nuevo rango y la columna Conductor añadida — sin esto Excel
    #    arroja "registros reparados" al abrir el archivo).
    saved_tables = []
    for tname in list(ws_resumen.tables):
        tbl = ws_resumen.tables[tname]
        saved_tables.append({
            'name': tname,
            'displayName': tbl.displayName,
            'tableStyleInfo': tbl.tableStyleInfo,
            'orig_cols': [{'id': c.id, 'name': c.name} for c in tbl.tableColumns],
        })
        del ws_resumen.tables[tname]

    # 1. Capturar filas originales ANTES del shift
    original_rows = []
    for r in range(RESUMEN_FIRST_DATA_ROW, ws_resumen.max_row + 1):
        alias = ws_resumen.cell(row=r, column=1).value
        if not alias:
            continue
        original_rows.append({
            'alias': str(alias).strip(),
            'values': [
                ws_resumen.cell(row=r, column=c).value
                for c in range(1, ws_resumen.max_column + 1)
            ],
        })

    last_row_pre = ws_resumen.max_row

    # 2. Insertar columna B (todo se desplaza una posición a la derecha)
    ws_resumen.insert_cols(2)

    # 3. Encabezado de Conductor
    ws_resumen.cell(row=RESUMEN_HEADER_ROW, column=2, value='Conductor')

    # 4. Borrar el área de datos (quedó shifteada en cols 1, 3..max+1)
    new_max_col = ws_resumen.max_column
    for r in range(RESUMEN_FIRST_DATA_ROW, last_row_pre + 1):
        for c in range(1, new_max_col + 1):
            ws_resumen.cell(row=r, column=c).value = None

    # 5. Reescribir: una fila por (vehículo, conductor)
    EXCESOS_COL_NEW = RESUMEN_COL_EXCESOS_OLD + 1   # 3
    VELMAX_COL_NEW = RESUMEN_COL_VELMAX_OLD + 1     # 7

    write_row = RESUMEN_FIRST_DATA_ROW
    for orig in original_rows:
        alias = orig['alias']
        conductors = by_vehicle.get(alias)

        if not conductors:
            # Sin entradas en Detalle: una fila con "Sin Conductor", preservando todo
            ws_resumen.cell(row=write_row, column=1, value=alias)
            ws_resumen.cell(row=write_row, column=2, value=SIN_CONDUCTOR)
            for i in range(1, len(orig['values'])):
                ws_resumen.cell(row=write_row, column=i + 2, value=orig['values'][i])
            write_row += 1
            continue

        if len(conductors) == 1:
            # Un solo conductor: preserva valores originales, solo añade nombre
            conductor = next(iter(conductors.keys()))
            ws_resumen.cell(row=write_row, column=1, value=alias)
            ws_resumen.cell(row=write_row, column=2, value=conductor)
            for i in range(1, len(orig['values'])):
                ws_resumen.cell(row=write_row, column=i + 2, value=orig['values'][i])
            write_row += 1
            continue

        # Múltiples conductores: una fila por conductor con métricas recalculadas
        for conductor, stats in conductors.items():
            ws_resumen.cell(row=write_row, column=1, value=alias)
            ws_resumen.cell(row=write_row, column=2, value=conductor)
            ws_resumen.cell(row=write_row, column=EXCESOS_COL_NEW, value=stats['excesos'])
            ws_resumen.cell(row=write_row, column=VELMAX_COL_NEW, value=stats['vel_max'])
            write_row += 1

    last_row = write_row - 1

    # 6. Recrear tabla(s) estructuradas con la nueva referencia y columnas
    for info in saved_tables:
        last_col_letter = get_column_letter(ws_resumen.max_column)
        new_ref = f'A{RESUMEN_HEADER_ROW}:{last_col_letter}{last_row}'
        new_cols = [TableColumn(id=1, name=info['orig_cols'][0]['name'])]  # Alias
        new_cols.append(TableColumn(id=2, name='Conductor'))               # nueva
        for i, oc in enumerate(info['orig_cols'][1:], start=3):
            new_cols.append(TableColumn(id=i, name=oc['name']))
        new_tbl = Table(displayName=info['displayName'], name=info['name'], ref=new_ref)
        new_tbl.tableColumns = new_cols
        if info['tableStyleInfo']:
            new_tbl.tableStyleInfo = info['tableStyleInfo']
        ws_resumen.add_table(new_tbl)

    return last_row


def make_bar_chart(ws, title, value_col, y_axis_title, style, last_row, cats_max_col=2):
    """Crea un BarChart de columnas.

    Si cats_max_col == 2 -> categorías multi-nivel (Vehículo + Conductor).
    Si cats_max_col == 1 -> categorías simples (solo Vehículo, modo v1).
    """
    ch = BarChart()
    ch.type = 'col'
    ch.style = style
    ch.title = title
    ch.x_axis.title = 'Unidad / Conductor' if cats_max_col == 2 else 'Unidad'
    ch.y_axis.title = y_axis_title
    ch.legend = None

    # Forzar visibilidad de ejes (workaround openpyxl)
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.x_axis.majorTickMark = 'out'
    ch.x_axis.minorTickMark = 'none'
    ch.y_axis.majorTickMark = 'out'
    ch.y_axis.minorTickMark = 'none'

    ch.width = 22
    ch.height = 13

    # Datos: la columna value_col, incluyendo el header (titles_from_data=True)
    data = Reference(
        ws,
        min_col=value_col, max_col=value_col,
        min_row=RESUMEN_HEADER_ROW, max_row=last_row,
    )
    cats = Reference(
        ws,
        min_col=1, max_col=cats_max_col,
        min_row=RESUMEN_FIRST_DATA_ROW, max_row=last_row,
    )
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    return ch


# ---------- Endpoints ----------
@app.route('/', methods=['GET'])
def health():
    return jsonify({
        'status': 'ok',
        'service': 'trackgts-excel-charts',
        'version': '2.0.0',
    })


@app.route('/add-charts', methods=['POST'])
def add_charts():
    if not check_auth():
        return jsonify({'error': 'Unauthorized — invalid or missing API key'}), 401
    try:
        data = request.get_json(silent=True)
        if not data or 'excelBase64' not in data:
            return jsonify({'error': 'Body debe incluir campo "excelBase64"'}), 400

        excel_bytes = base64.b64decode(data['excelBase64'])
        wb = load_workbook(io.BytesIO(excel_bytes))

        resumen_name = find_sheet(wb, 'resumen')
        detalle_name = find_sheet(wb, 'detalle')
        if not resumen_name:
            return jsonify({
                'error': 'No se encontró hoja "Resumen"',
                'sheetsFound': wb.sheetnames,
            }), 400
        if not detalle_name:
            return jsonify({
                'error': 'No se encontró hoja "Detalle"',
                'sheetsFound': wb.sheetnames,
            }), 400

        ws_resumen = wb[resumen_name]
        ws_detalle = wb[detalle_name]

        # Detectar si Detalle 1 incluye columna Conductor (saved-report v2 vs API v1)
        detalle_header_row = DETALLE_FIRST_DATA_ROW - 1
        conductor_col = find_column_by_header(ws_detalle, detalle_header_row, 'conductor')

        if conductor_col is None:
            # ===== MODO v1 FALLBACK: sin columna Conductor =====
            # El Excel viene del API básico sin Conductor. Generamos los gráficos
            # como en la versión 1.0 (Vehículo vs Excesos / Vel Máxima) sin tocar datos.
            log.info('Detalle 1 NO tiene columna Conductor — usando modo v1 (sin Conductor)')

            last_row = ws_resumen.max_row
            for r in range(last_row, RESUMEN_HEADER_ROW, -1):
                if ws_resumen.cell(row=r, column=1).value:
                    last_row = r
                    break
            if last_row < RESUMEN_FIRST_DATA_ROW:
                return jsonify({'error': 'No hay filas de datos en Resumen'}), 400
            rows_count = last_row - RESUMEN_HEADER_ROW
            chart_anchor_row = last_row + 2

            chart1 = make_bar_chart(
                ws_resumen, 'Excesos de Velocidad', RESUMEN_COL_EXCESOS_OLD,
                'Cantidad de Excesos', 10, last_row, cats_max_col=1,
            )
            ws_resumen.add_chart(chart1, f'A{chart_anchor_row}')

            chart2 = make_bar_chart(
                ws_resumen, 'Velocidad Máxima', RESUMEN_COL_VELMAX_OLD,
                'Velocidad Máxima (Km/h)', 11, last_row, cats_max_col=1,
            )
            ws_resumen.add_chart(chart2, f'F{chart_anchor_row}')

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            result_b64 = base64.b64encode(output.read()).decode('utf-8')

            return jsonify({
                'excelBase64': result_b64,
                'info': {
                    'sheet': resumen_name,
                    'rowsCount': rows_count,
                    'chartAnchorRow': chart_anchor_row,
                    'mode': 'v1-fallback',
                    'note': 'Detalle 1 no tiene columna Conductor — gráficos sin desglose por conductor',
                },
            })

        # ===== MODO v2: Detalle 1 tiene columna Conductor =====
        log.info(f'Conductor en columna {conductor_col} de Detalle — modo v2 (con Conductor)')

        # Override del índice global por si el header no está en la columna 2 estándar
        global DETALLE_COL_CONDUCTOR
        DETALLE_COL_CONDUCTOR = conductor_col

        # 1. Agregar Detalle por (vehículo, conductor)
        by_vehicle = aggregate_detalle(ws_detalle)
        log.info(f'Detalle agregado: {len(by_vehicle)} vehículos únicos')

        # 2. Insertar columna Conductor y desglosar filas
        last_row = rebuild_resumen_with_conductor(ws_resumen, by_vehicle)
        if last_row < RESUMEN_FIRST_DATA_ROW:
            return jsonify({'error': 'No hay filas de datos en Resumen'}), 400
        rows_count = last_row - RESUMEN_FIRST_DATA_ROW + 1
        log.info(f'Resumen reconstruido con {rows_count} filas (vehículo+conductor)')

        # 3. Construir gráficos con categorías multi-nivel
        EXCESOS_COL = 3   # columna C tras el shift
        VELMAX_COL = 7    # columna G tras el shift
        chart_anchor_row = last_row + 2

        chart1 = make_bar_chart(
            ws_resumen, 'Excesos de Velocidad', EXCESOS_COL,
            'Cantidad de Excesos', 10, last_row, cats_max_col=2,
        )
        ws_resumen.add_chart(chart1, f'A{chart_anchor_row}')

        chart2 = make_bar_chart(
            ws_resumen, 'Velocidad Máxima', VELMAX_COL,
            'Velocidad Máxima (Km/h)', 11, last_row, cats_max_col=2,
        )
        # Anclar chart2 más a la derecha para no solapar con chart1 (dato shifteado)
        ws_resumen.add_chart(chart2, f'H{chart_anchor_row}')

        # 4. Devolver Excel modificado
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        result_b64 = base64.b64encode(output.read()).decode('utf-8')

        return jsonify({
            'excelBase64': result_b64,
            'info': {
                'sheet': resumen_name,
                'rowsCount': rows_count,
                'chartAnchorRow': chart_anchor_row,
                'vehiclesProcessed': len(by_vehicle),
                'mode': 'v2-with-conductor',
            },
        })
    except Exception as e:
        log.exception('Error procesando Excel')
        return jsonify({'error': str(e), 'type': type(e).__name__}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
